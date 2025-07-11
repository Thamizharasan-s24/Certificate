from flask import Flask, render_template, request, send_file, redirect, url_for
from PIL import Image, ImageDraw, ImageFont
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
import io

app = Flask(__name__)
preview_cache = {}

# Function to log data into Excel
def log_to_excel(name, role, start_date, end_date, filename):
    excel_path = "data/certificates.xlsx"
    headers = ["Name", "Role", "Start Date", "End Date", "Filename", "Generated Time"]
    now = datetime.now().strftime("%d %b %Y, %I:%M %p")

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    ws.append([name, role, start_date, end_date, filename, now])
    wb.save(excel_path)

@app.route('/')
def index():
    return render_template("index.html")

@app.route('/generate', methods=['POST'])
def generate():
    name = request.form['name']
    role = request.form['role']
    start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').strftime('%d %b %Y')
    end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d').strftime('%d %b %Y')
    file_format = request.form['file_format']
    template_file = request.files['template']

    # Load image template
    image = Image.open(template_file.stream).convert("RGB")
    draw = ImageDraw.Draw(image)

    # Load fonts
    name_font = ImageFont.truetype("static/fonts/GreatVibes-Regular.ttf", 90)
    poppins_font = ImageFont.truetype("static/fonts/Poppins-Regular.ttf", 26)
    poppins_bold = ImageFont.truetype("static/fonts/Poppins-Bold.ttf", 26)

    # Draw name
    name_width = draw.textlength(name, font=name_font)
    name_x = (image.width - name_width) // 2
    draw.text((name_x, 380), name, fill=(255, 69, 0), font=name_font)  # red-orange

    # Paragraph content
    paragraph_parts = [
        ("This is to certify that the ", poppins_font),
        (role.upper(), poppins_bold),
        (" Internship at ", poppins_font),
        ("DiyanTech Solutions Pvt Ltd", poppins_bold),
        (" was successfully completed over a duration of one month, from ", poppins_font),
        (start_date, poppins_bold),
        (" to ", poppins_font),
        (end_date, poppins_bold),
        (".", poppins_font)
    ]

    def get_line_width(parts):
        return sum(draw.textlength(text, font=font) for text, font in parts)

    def draw_line(line_parts, y_pos):
        x = (image.width - get_line_width(line_parts)) // 2
        for text, font in line_parts:
            draw.text((x, y_pos), text, font=font, fill="black")
            x += draw.textlength(text, font=font)

    # Wrap and draw text
    y = 520
    max_width = 1000
    line_spacing = 40
    current_line = []
    current_width = 0

    for text, font in paragraph_parts:
        part_words = text.split(" ")
        for i, word in enumerate(part_words):
            spacing = " " if i != len(part_words) - 1 else ""
            word_text = word + spacing
            word_width = draw.textlength(word_text, font=font)
            if current_width + word_width > max_width:
                draw_line(current_line, y)
                y += line_spacing
                current_line = [(word_text, font)]
                current_width = word_width
            else:
                current_line.append((word_text, font))
                current_width += word_width

    if current_line:
        draw_line(current_line, y)

    # Save image in memory
    img_io = io.BytesIO()
    format_map = {'png': 'PNG', 'jpg': 'JPEG', 'pdf': 'PDF'}
    image_format = format_map[file_format]
    image.save(img_io, format=image_format)
    img_io.seek(0)

    filename = f"{name.replace(' ', '_')}_certificate.{file_format}"
    preview_cache['image'] = img_io.getvalue()
    preview_cache['format'] = file_format
    preview_cache['filename'] = filename

    # Log details into Excel
    log_to_excel(name, role, start_date, end_date, filename)

    return render_template("preview.html", filename=filename)

@app.route('/download')
def download():
    if 'image' in preview_cache:
        return send_file(
            io.BytesIO(preview_cache['image']),
            mimetype=f"image/{preview_cache['format']}",
            as_attachment=True,
            download_name=preview_cache['filename']
        )
    return redirect(url_for('index'))

@app.route('/cancel')
def cancel():
    preview_cache.clear()
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=80)
